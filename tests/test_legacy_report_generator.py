from pathlib import Path

import pandas as pd
from docx import Document
from openpyxl import load_workbook

from utils.report_generator import ReportGenerator


def _sample_processed_data() -> dict:
    return {
        "centers": pd.DataFrame(
            [
                {"z": 5.0, "x": 0.0, "y": 0.0, "deviation": 0.001},
                {"z": 15.0, "x": 0.0, "y": 0.0, "deviation": 0.002},
            ]
        ),
        "straightness_profiles": [
            {
                "part_number": 1,
                "belt": 1,
                "part_min_height": 0.0,
                "part_max_height": 10.0,
                "points": [
                    {"z": 0.0, "deflection_mm": 0.0, "tolerance_mm": 13.3, "section_length_m": 10.0},
                    {"z": 5.0, "deflection_mm": 12.0, "tolerance_mm": 13.3, "section_length_m": 10.0},
                    {"z": 10.0, "deflection_mm": 0.0, "tolerance_mm": 13.3, "section_length_m": 10.0},
                ],
            },
            {
                "part_number": 2,
                "belt": 2,
                "part_min_height": 10.0,
                "part_max_height": 20.0,
                "points": [
                    {"z": 10.0, "deflection_mm": 0.0, "tolerance_mm": 13.3, "section_length_m": 10.0},
                    {"z": 15.0, "deflection_mm": 18.0, "tolerance_mm": 13.3, "section_length_m": 10.0},
                    {"z": 20.0, "deflection_mm": 0.0, "tolerance_mm": 13.3, "section_length_m": 10.0},
                ],
            },
        ],
    }


def test_generate_excel_report_uses_canonical_part_aware_straightness(tmp_path: Path):
    output = tmp_path / "legacy.xlsx"

    ReportGenerator().generate_excel_report(pd.DataFrame(), _sample_processed_data(), str(output))

    workbook = load_workbook(output)
    ws_results = workbook[workbook.sheetnames[0]]
    ws_normatives = workbook[workbook.sheetnames[1]]
    ws_straightness = workbook[workbook.sheetnames[2]]

    assert workbook.sheetnames == ["Результаты", "Нормативы", "Прямолинейность"]
    assert ws_results["F5"].value == 12
    assert ws_results["F6"].value == 18
    assert ws_straightness["A1"].value == "Расчет стрелы прогиба поясов"
    assert ws_straightness["A8"].value == 2
    assert ws_straightness["B8"].value == 2
    assert ws_straightness["F8"].value == "Превышение"
    assert "Максимальная стрела прогиба" in str(ws_normatives["A18"].value)


def test_generate_docx_report_uses_canonical_part_aware_straightness(tmp_path: Path):
    output = tmp_path / "legacy.docx"

    ReportGenerator().generate_docx_report(
        pd.DataFrame(),
        _sample_processed_data(),
        str(output),
        object_info={"executor": "QA"},
    )

    document = Document(output)
    text = "\n".join(paragraph.text for paragraph in document.paragraphs if paragraph.text.strip())

    assert "РАСЧЕТ СТРЕЛЫ ПРОГИБА ПОЯСА СТВОЛА" in text
    assert "Превышений: 1 из 6 точек." in text
    assert "часть 2, пояс 2, высота 15.000 м" in text
    assert "Ð" not in text
